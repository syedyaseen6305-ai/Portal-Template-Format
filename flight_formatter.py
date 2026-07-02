import streamlit as st
import pandas as pd
from datetime import datetime, time
import io
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.title("✈️ Portal Data Formatter (with Template Paste)")

# Final smart rollover logic
def format_datetime(date, raw_time, base_time=None):
    if pd.isna(date) or pd.isna(raw_time):
        return None
    try:
        base_date = pd.to_datetime(date).date()

        def to_time(val):
            if isinstance(val, str):
                return datetime.strptime(val.strip(), "%H:%M").time()
            elif isinstance(val, time):
                return val
            else:
                return None

        raw_time_obj = to_time(raw_time)
        base_time_obj = to_time(base_time)

        if base_time_obj and base_time_obj >= time(18, 0) and raw_time_obj < time(3, 0):
            base_date += pd.Timedelta(days=1)

        full_datetime = datetime.combine(base_date, raw_time_obj.replace(second=0))
        return full_datetime
    except Exception:
        return None

def extract_services(row):
    services = []
    for col in row.index:
        if isinstance(col, str) and str(row[col]).strip() == '√':
            services.append(col.strip())

    remark = str(row.get('OTHER SERVICES/REMARKS', '')).upper()
    if 'ON CALL - NEEDED ENGINEER SUPPORT' in remark:
        services.append('On Call')
    elif 'CANCELED WITHOUT NOTICE' in remark or 'CANCELLED WITHOUT NOTICE' in remark:
        services.append('Canceled without notice')
    elif 'CANCELED' in remark or 'CANCELLED' in remark:
        services.append('Cancelled Flight')
    elif 'ON CALL' in remark:
        services.append('Per Landing')

    corrected_services = []
    for service in services:
        if service == 'TECH. SUPT':
            corrected_services.append('TECH SUPPORT')
        elif service == 'HEAD SET':
            corrected_services.append('Headset')
        else:
            corrected_services.append(service)

    return ', '.join(corrected_services) if corrected_services else None

def categorize(row):
    remark = str(row.get('OTHER SERVICES/REMARKS', '')).upper()
    if 'TRANSIT' in remark:
        return '1_TRANSIT'
    elif 'ON CALL - NEEDED ENGINEER SUPPORT' in remark:
        return '2_ONCALL_ENGINEER'
    elif 'CANCELED WITHOUT NOTICE' in remark or 'CANCELLED WITHOUT NOTICE' in remark:
        return '3_CANCELED'
    elif 'ON CALL' in remark:
        return '4_ONCALL_RECORDED'
    else:
        return '5_OTHER'

def process_file(uploaded_file, template_file):
    df = pd.read_excel(uploaded_file, sheet_name='Daily Operations Report', header=4)
    df.dropna(how='all', inplace=True)
    df.rename(columns=lambda x: x.strip() if isinstance(x, str) else x, inplace=True)
    df.rename(columns={
        'REG.': 'REG',
        'TECH.\nSUPT': 'TECH. SUPT',
        'TECH. SUPT': 'TECH SUPPORT',
        'HEAD SET': 'Headset',
        'TRANSIT': 'Transit',
        'WKLY CK': 'Weekly Check',
        'DAILY CK': 'Daily Check'
    }, inplace=True)

    df['STA.'] = df.apply(lambda row: format_datetime(row['DATE'], row['STA'], None), axis=1)
    df['ATA.'] = df.apply(lambda row: format_datetime(row['DATE'], row['ATA'], row['STA']), axis=1)
    df['STD.'] = df.apply(lambda row: format_datetime(row['DATE'], row.get('STD'), row['STA']), axis=1)
    df['ATD.'] = df.apply(lambda row: format_datetime(row['DATE'], row.get('ATD'), row['STA']), axis=1)

    canceled_mask = df['OTHER SERVICES/REMARKS'].str.contains('CANCELED|CANCELLED', case=False, na=False)
    df.loc[canceled_mask, 'ATA.'] = df.loc[canceled_mask, 'STA.']
    df.loc[canceled_mask, 'ATD.'] = df.loc[canceled_mask, 'STD.']

    df['Customer'] = df['FLT NO.'].astype(str).str.strip().apply(lambda x: 'XLR' if x.startswith('DHX') else x[:2])
    df['Services'] = df.apply(extract_services, axis=1)
    df['Is Canceled'] = df['OTHER SERVICES/REMARKS'].str.contains('CANCELED|CANCELLED', na=False, case=False)
    df['Category'] = df.apply(categorize, axis=1)
    df.sort_values(by=['Category', 'STA.'], inplace=True)

    result_rows = []
    for _, row in df.iterrows():
        try:
            result_rows.append({
                'WO#': row['W/O'],
                'Station': 'KKIA',
                'Customer': row['Customer'],
                'Flight No.': row['FLT NO.'],
                'Registration Code': row['REG'],
                'Aircraft': row['A/C TYPES'],
                'Date': pd.to_datetime(row['DATE']),
                'STA.': row['STA.'],
                'ATA.': row['ATA.'],
                'STD.': row['STD.'],
                'ATD.': row['ATD.'],
                'Is Canceled': row['Is Canceled'],
                'Services': row['Services'],
                'Employees': ', '.join(filter(None, [
                    str(int(row['ENGR'])) if pd.notna(row['ENGR']) and str(row['ENGR']).replace('.', '', 1).isdigit() else '',
                    str(int(row['TECH'])) if pd.notna(row['TECH']) and str(row['TECH']).replace('.', '', 1).isdigit() else ''
                ])),
                'Remarks': '',
                'Comments': ''
            })
        except Exception:
            pass

    result_df = pd.DataFrame(result_rows)

    output = io.BytesIO()
    template_wb = load_workbook(template_file)
    ws = template_wb.active

    # Write starting after the header row (assumed row 2)
    start_row = 2
    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value

            # Apply short date format ONLY to 'Date' column (column 7)
            if c_idx == 7 and isinstance(value, (datetime, pd.Timestamp)):
                cell.value = value.date()  # Remove time part
                cell.number_format = 'yyyy/mm/dd'
            elif isinstance(value, pd.Timestamp):
                cell.number_format = 'yyyy/mm/dd hh:mm'

    template_wb.save(output)
    output.seek(0)

    report_date = df['DATE'].iloc[0] if not df.empty else None
    return output, report_date

# Upload files
uploaded_file = st.file_uploader("Upload Daily Operations Report", type=["xlsx"])
template_file = st.file_uploader("Upload Work Order Template", type=["xlsx"])

if uploaded_file and template_file:
    st.success("✅ Files uploaded successfully!")
    final_output, report_date = process_file(uploaded_file, template_file)

    if report_date is not None:
        try:
            date_obj = pd.to_datetime(report_date)
            filename = date_obj.strftime("%d%b%Y").upper() + "_WorkOrders.xlsx"
        except Exception:
            filename = "Final_WorkOrders.xlsx"
    else:
        filename = "Final_WorkOrders.xlsx"

    st.download_button("📥 Download Final Work Order File", data=final_output, file_name=filename)
